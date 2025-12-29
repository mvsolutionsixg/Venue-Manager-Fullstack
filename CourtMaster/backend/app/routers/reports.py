from fastapi import APIRouter, Depends, HTTPException, status
from sqlalchemy.orm import Session
from fastapi.responses import StreamingResponse
from typing import Optional
from datetime import date
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill

from ..database import get_db
from .. import models, crud

router = APIRouter(
    tags=["reports"],
)

@router.get("/bookings/export")
def export_bookings(
    date: Optional[date] = None,
    search: Optional[str] = None,
    db: Session = Depends(get_db)
):
    # Reuse existing filter logic
    bookings = crud.get_bookings(db, skip=0, limit=10000, target_date=date, search=search)
    
    # Create Excel Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bookings Report"
    
    # Headers
    headers = ["ID", "Customer Name", "Mobile", "Date", "Start Time", "End Time", "Court ID", "Status"]
    ws.append(headers)
    
    # Style Headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid") # Indigo Primary
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        
    # Data
    for booking in bookings:
        ws.append([
            booking.id,
            booking.customer_name,
            booking.mobile,
            booking.date,
            booking.start_time,
            booking.end_time,
            booking.court_id,
            booking.status
        ])
        
    # Auto-adjust column width (approximate)
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"bookings_export_{date if date else 'all'}.xlsx"
    
    headers = {
        'Content-Disposition': f'attachment; filename="{filename}"'
    }
    
    return StreamingResponse(
        buffer, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        headers=headers
    )

@router.get("/dashboard/stats")
def dashboard_stats(period: str = "overall", db: Session = Depends(get_db)):
    return crud.get_dashboard_stats(db, period=period)

@router.get("/dashboard/charts")
def dashboard_charts(db: Session = Depends(get_db)):
    daily = crud.get_daily_bookings_chart(db)
    status_dist = crud.get_booking_status_distribution(db)
    return {"daily": daily, "status": status_dist}

@router.get("/capacity")
def capacity_heatmap(
    start_date: date,
    end_date: date,
    db: Session = Depends(get_db)
):
    return crud.get_court_capacity_heatmap(db, start_date, end_date)
